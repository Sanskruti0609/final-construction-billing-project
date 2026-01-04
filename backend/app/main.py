# app/main.py
from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse

import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from textwrap import wrap
import re

from openpyxl import Workbook  # for Excel export

from .database import SessionLocal, engine
from . import models, schemas, crud
from .utils.ssr_loader import fetch_ssr_rate
from .utils.boq_loader import fetch_boq_item_no   # <--- NEW IMPORT
from fastapi import Request

models.Base.metadata.create_all(bind=engine)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],           # dev mode
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/")
def root():
    return {"message": "Construction Billing API is running"}


# ---------- SSR preview (for rate popup in form) ----------
@app.post("/ssr/rate", response_model=schemas.RateResponse)
def preview_rate(req: schemas.RateRequest):
    """
    Strict SSR + BOQ behaviour:

    - Try STRICT SSR match:
        - If found → return SSR item no, unit, rates, gst, amount, and BOQ item no (if exists).
    - If NOT found in SSR:
        - Try BOQ match:
            - If BOQ description found → treat as NON SSR ITEM, return:
                * ssr_item_no = "NON SSR ITEM"
                * unit = "" (user will enter manually)
                * rates = 0 (user will enter)
                * boq_item_no = from BOQ
                * non_ssr = True
            - If not even in BOQ → raise 404.
    """

    ssr_info = fetch_ssr_rate(req.description, req.quantity)
    boq_no = fetch_boq_item_no(req.description)

    # ---------- CASE 1: SSR FOUND (exact match) ----------
    if ssr_info is not None:
        return {
            **ssr_info,
            "boq_item_no": boq_no,
            "non_ssr": False,
        }

    # ---------- CASE 2: SSR NOT FOUND, BUT BOQ EXISTS ----------
    if boq_no:
        # NON SSR item which still exists in BOQ file
        return {
            "ssr_item_no": "NON SSR ITEM",
            "unit": "",
            "base_rate": 0.0,
            "gst_rate": 0.0,
            "final_rate": 0.0,
            "total_amount": 0.0,
            "boq_item_no": boq_no,
            "non_ssr": True,
        }

    # ---------- CASE 3: Neither SSR nor BOQ have this description ----------
    raise HTTPException(
        status_code=404,
        detail="Item not found in SSR or BOQ for this description",
    )



# ---------- MATERIALS CRUD ----------
@app.get("/materials/", response_model=list[schemas.Material])
def list_materials(db: Session = Depends(get_db)):
    return crud.get_materials(db)

@app.post("/materials/", response_model=schemas.Material)
def create_material(material: schemas.MaterialCreate, db: Session = Depends(get_db)):
    # material.quantity should already be total (sum of all L*B*D*No) from frontend
    return crud.create_material(db, material)

# @app.post("/materials/")
# async def create_material(request: Request, db: Session = Depends(get_db)):
#     body = await request.json()
#     print("📦 RAW REQUEST BODY:", body)
#     return {"received": body}


@app.delete("/materials/{material_id}")
def remove_material(material_id: int, db: Session = Depends(get_db)):
    ok = crud.delete_material(db, material_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Material not found")
    return {"success": True}


# ============================================================
#  FULL MATERIALS BILL (ALL ITEMS) - PDF
# ============================================================
@app.get("/materials/bill/pdf")
def download_materials_bill(db: Session = Depends(get_db)):
    materials = crud.get_materials(db)
    if not materials:
        raise HTTPException(status_code=400, detail="No materials to include in bill")

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    margin_left = 30
    margin_right = width - 30
    table_width = margin_right - margin_left

    # -------- COLUMN POSITIONS (11 vertical lines = 10 columns) ----------
    col1_x  = margin_left
    col2_x  = col1_x + 35   # 1
    col3_x  = col2_x + 35   # 2
    col4_x  = col3_x + 55   # 3
    col5_x  = col4_x + 75   # 4 (Quantity)
    col6_x  = col5_x + 210  # 5 (Items of work)
    col7_x  = col6_x + 60   # 6 (Rate)
    col8_x  = col7_x + 70   # 7 (Unit)
    col9_x  = col8_x + 70   # 8 (Amt Up-to-date)
    col10_x = col9_x + 60   # 9 (Amt Since previous bill)
    last_x  = margin_right  # 10 (Remarks)

    line_h = 10

    def ensure_space(h_needed: float):
        """Start new page if not enough space for h_needed."""
        nonlocal y
        if y - h_needed < 110:
            p.showPage()
            draw_header()

    def draw_header():
        nonlocal y
        p.setFont("Helvetica-Bold", 11)
        y = height - 55
        p.drawCentredString(width / 2, y, "Part I - Account of work executed")
        y -= 8

        header_top = y
        header_bottom = header_top - (3 * line_h)  # 3 text rows
        row2_y = header_top - line_h
        row3_y = header_top - 2 * line_h

        # outer rectangle
        p.setLineWidth(0.5)
        p.rect(margin_left, header_bottom, table_width, header_top - header_bottom)

        # vertical lines
        for x in [col1_x, col2_x, col3_x, col4_x, col5_x,
                  col6_x, col7_x, col8_x, col9_x, col10_x, last_x]:
            p.line(x, header_bottom, x, header_top)

        # horizontal lines to split 3 header rows
        p.line(margin_left, row2_y, margin_right, row2_y)
        p.line(margin_left, row3_y, margin_right, row3_y)

        # ---------- HEADER TEXT ----------
        p.setFont("Helvetica-Bold", 7)

        # Row 1 – grouped headings
        # cols 1–3 : Advance payments...
        p.drawCentredString(
            (col1_x + col4_x) / 2,
            header_top - 3,
            "Advance payments, for work"
        )
        p.drawCentredString(
            (col1_x + col4_x) / 2,
            header_top - 3 - line_h,
            "done not yet measured"
        )

        # col 4 : Quantity executed…
        p.drawCentredString(
            (col4_x + col5_x) / 2,
            header_top - 3,
            "Quantity"
        )
        p.drawCentredString(
            (col4_x + col5_x) / 2,
            header_top - 3 - line_h,
            "executed up-"
        )
        p.drawCentredString(
            (col4_x + col5_x) / 2,
            header_top - 3 - 2 * line_h,
            "to date as per"
        )
        p.drawCentredString(
            (col4_x + col5_x) / 2,
            header_top - 3 - 3 * line_h,
            "measurement book"
        )

        # col 5 : Items of work
        p.drawCentredString(
            (col5_x + col6_x) / 2,
            header_top - 3,
            "Items of work"
        )
        p.drawCentredString(
            (col5_x + col6_x) / 2,
            header_top - 3 - line_h,
            "(Grouped Under Sub-heads or Sub-works of estimate)"
        )

        # col 6 : Rate
        p.drawCentredString(
            (col6_x + col7_x) / 2,
            header_top - 3,
            "Rate"
        )
        p.drawCentredString(
            (col6_x + col7_x) / 2,
            header_top - 3 - line_h,
            "(Total fees payable)"
        )

        # col 7 : Unit
        p.drawCentredString(
            (col7_x + col8_x) / 2,
            header_top - 3,
            "Unit"
        )

        # cols 8–9 : Payment on the basis...
        p.drawCentredString(
            (col8_x + col10_x) / 2,
            header_top - 3,
            "Payment on the basis of actual"
        )
        p.drawCentredString(
            (col8_x + col10_x) / 2,
            header_top - 3 - line_h,
            "measurements"
        )

        # col 10 : Remarks...
        p.drawCentredString(
            (col10_x + last_x) / 2,
            header_top - 3,
            "Remarks"
        )
        p.drawCentredString(
            (col10_x + last_x) / 2,
            header_top - 3 - line_h,
            "(with reasons for delay in"
        )
        p.drawCentredString(
            (col10_x + last_x) / 2,
            header_top - 3 - 2 * line_h,
            "adjusting payments shown"
        )
        p.drawCentredString(
            (col10_x + last_x) / 2,
            header_top - 3 - 3 * line_h,
            "in column (1)"
        )

        # Row 2 – sub-headings for cols 1–3 and 8–9
        p.setFont("Helvetica-Bold", 7)

        # col1
        p.drawCentredString(
            (col1_x + col2_x) / 2,
            row3_y + 3,
            "Total as"
        )
        p.drawCentredString(
            (col1_x + col2_x) / 2,
            row3_y + 3 - line_h,
            "per previous"
        )
        p.drawCentredString(
            (col1_x + col2_x) / 2,
            row3_y + 3 - 2 * line_h,
            "bill"
        )

        # col2
        p.drawCentredString(
            (col2_x + col3_x) / 2,
            row3_y + 3,
            "Since"
        )
        p.drawCentredString(
            (col2_x + col3_x) / 2,
            row3_y + 3 - line_h,
            "previous"
        )
        p.drawCentredString(
            (col2_x + col3_x) / 2,
            row3_y + 3 - 2 * line_h,
            "bill"
        )

        # col3
        p.drawCentredString(
            (col3_x + col4_x) / 2,
            row3_y + 3,
            "Total up-"
        )
        p.drawCentredString(
            (col3_x + col4_x) / 2,
            row3_y + 3 - line_h,
            "to date"
        )

        # col8 / 9
        p.drawCentredString(
            (col8_x + col9_x) / 2,
            row3_y + 3,
            "Up-to-date"
        )
        p.drawCentredString(
            (col9_x + col10_x) / 2,
            row3_y + 3,
            "Since previous bill"
        )

        # Row 3 – column numbers 1..10
        num_y = header_bottom + 2
        for idx, (xl, xr) in enumerate(
            [
                (col1_x, col2_x),
                (col2_x, col3_x),
                (col3_x, col4_x),
                (col4_x, col5_x),
                (col5_x, col6_x),
                (col6_x, col7_x),
                (col7_x, col8_x),
                (col8_x, col9_x),
                (col9_x, col10_x),
                (col10_x, last_x),
            ],
            start=1,
        ):
            p.drawCentredString((xl + xr) / 2, num_y, str(idx))

        y = header_bottom - 14
        p.setFont("Helvetica", 8)

    # -------- START FIRST PAGE --------
    draw_header()

    grand_without_18 = 0.0

    for m in materials:
        qty = float(m.quantity or 0.0)
        amount = float(m.total_amount or 0.0)
        base_rate = float(m.base_rate or 0.0)
        unit = (m.unit or "").strip()
        # use saved BOQ item no; fallback "-" if not present
        boq_no = (getattr(m, "boq_item_no", "") or "").strip() or "-"

        # clean + wrap description
        raw_desc = m.description or ""
        clean_desc = re.sub(r"\s+", " ", raw_desc.replace("\n", " ").replace("\r", " ")).strip()
        desc_lines = wrap(clean_desc, 90) or [""]

        # height = one top line + all desc lines + padding
        row_height = line_h * (1 + len(desc_lines)) + 10

        ensure_space(row_height + 4)

        top_y = y
        bottom_y = y - row_height

        # row rectangle & verticals
        p.rect(margin_left, bottom_y, table_width, row_height)
        for x in [col1_x, col2_x, col3_x, col4_x, col5_x,
                  col6_x, col7_x, col8_x, col9_x, col10_x, last_x]:
            p.line(x, bottom_y, x, top_y)

        # text baseline a bit lower to avoid touching borders
        text_y = top_y - 6

        # column 4 → Quantity
        p.drawRightString(col5_x - 3, text_y, f"{qty:.3f}")

        # column 5 → "Item No. X" and description
        p.setFont("Helvetica-Bold", 8)
        p.drawString(col5_x + 2, text_y, f"Item No. {boq_no}")
        p.setFont("Helvetica", 8)
        desc_y = text_y - line_h
        for line in desc_lines:
            p.drawString(col5_x + 2, desc_y, line)
            desc_y -= line_h

        # column 6 → Rate
        p.drawRightString(col6_x + 55, text_y, f"{base_rate:.2f}")
        # column 7 → Unit
        p.drawString(col7_x + 2, text_y, unit)
        # columns 8 & 9 → Amount (same)
        p.drawRightString(col8_x + 65, text_y, f"{amount:.2f}")
        p.drawRightString(col9_x + 55, text_y, f"{amount:.2f}")
        # column 10 → Remarks blank

        grand_without_18 += amount
        # small gap between rows
        y = bottom_y - 4

    # ---------- TOTALS + 18% GST ----------
    ensure_space(7 * line_h + 30)

    gst_18 = round(grand_without_18 * 0.18, 2)
    total_with_18 = round(grand_without_18 + gst_18, 2)

    def total_row(label: str, val: float, bold: bool = False):
        nonlocal y
        p.setFont("Helvetica-Bold" if bold else "Helvetica", 8)
        p.drawString(col5_x + 2, y, label)
        p.drawRightString(col8_x + 65, y, f"{val:.2f}")
        p.drawRightString(col9_x + 55, y, f"{val:.2f}")
        y -= line_h

    total_row("A)", grand_without_18, bold=True)
    total_row("(-)", 0.0)
    total_row("", grand_without_18)
    total_row("18% GST", gst_18)
    total_row("Total", total_with_18, bold=True)
    total_row("Price Escallation", 0.0)
    total_row("Grand Total", total_with_18, bold=True)

    # ---------- SIGNATURE BLOCK ----------
    if y < 90:
        p.showPage()
        y = height - 100

    p.setFont("Helvetica", 9)
    p.drawString(margin_left, 70, "Deputy Engineer,")
    p.drawString(margin_left, 58, "Bandra (P.W.) Project Sub Division No. 2")
    p.drawString(margin_left, 46, "Bandra")

    p.drawString(width - 260, 70, "Executive Engineer,")
    p.drawString(width - 260, 58, "North Mumbai Division,")
    p.drawString(width - 260, 46, "Andheri, Mumbai.")

    p.showPage()
    p.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=materials_bill.pdf"},
    )

# ============================================================
#  FULL MATERIALS BILL (ALL ITEMS) - EXCEL
# ============================================================
@app.get("/materials/bill/excel")
def download_materials_bill_excel(db: Session = Depends(get_db)):
    materials = crud.get_materials(db)
    if not materials:
        raise HTTPException(status_code=400, detail="No materials to include in bill")

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Bill"

    # Header row (10 columns)
    # 1–3 blank, 4=Qty, 5=Item No + Desc, 6=Rate, 7=Unit, 8=Amount, 9=Amount, 10 blank
    ws.append([
        "", "", "",              # 1,2,3 blank
        "Quantity",              # 4
        "Items of work (Item No + Description)",  # 5
        "Rate",                  # 6
        "Unit",                  # 7
        "Amount (Col 8)",        # 8
        "Amount (Col 9)",        # 9
        ""                       # 10 blank
    ])

    grand_without_18 = 0.0

    for m in materials:
        qty = float(m.quantity or 0.0)
        amount = float(m.total_amount or 0.0)
        base_rate = float(m.base_rate or 0.0)
        unit = m.unit or ""
        boq_no = getattr(m, "boq_item_no", "") or ""

        desc = m.description or ""
        item_label = f"Item No. {boq_no}".strip() if boq_no else "Item"
        item_text = f"{item_label} - {desc}" if desc else item_label

        grand_without_18 += amount

        ws.append([
            "", "", "",              # col1,2,3
            qty,                     # col4
            item_text,               # col5
            base_rate,               # col6
            unit,                    # col7
            amount,                  # col8
            amount,                  # col9
            ""                       # col10
        ])

    # Totals section at the bottom (similar to PDF)
    gst_18 = round(grand_without_18 * 0.18, 2)
    total_with_18 = round(grand_without_18 + gst_18, 2)

    ws.append([])  # blank row
    ws.append(["", "", "", "", "A)", "", "", grand_without_18, grand_without_18, ""])
    ws.append(["", "", "", "", "(-)", "", "", 0.0, 0.0, ""])
    ws.append(["", "", "", "", "", "", "", grand_without_18, grand_without_18, ""])
    ws.append(["", "", "", "", "18%", "", "", gst_18, gst_18, ""])
    ws.append(["", "", "", "", "Total", "", "", total_with_18, total_with_18, ""])
    ws.append(["", "", "", "", "Price Escallation", "", "", 0.0, 0.0, ""])
    ws.append(["", "", "", "", "Grand Total", "", "", total_with_18, total_with_18, ""])

    # Set number format for numeric columns (4,6,8,9)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=9):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materials_bill.xlsx"},
    )

# ============================================================
#  SINGLE MATERIAL MEASUREMENT SHEET - PDF (SSR + BOQ + NON-SSR)
# ============================================================
@app.post("/materials/single-bill/pdf")
def download_single_material_bill(req: schemas.MaterialSingleBillRequest):
    """
    Measurement-style PDF for ONE item with multiple rows:
    Sr., Pile Description, No, B, D, L, Unit, Quantity

    - SSR Item No + BOQ Item No printed at top.
    - If SSR not found -> label 'NON SSR ITEM' and no unit.
    """

    # ---------- Normalize entries ----------
    entries = []
    if getattr(req, "entries", None):
        entries = list(req.entries)

    if not entries and any(
        v is not None
        for v in [req.no_of_items, req.length, req.breadth, req.depth, req.quantity]
    ):
        entries.append(
            schemas.MaterialMeasurementEntry(
                pile_description=None,
                no_of_items=req.no_of_items,
                length=req.length,
                breadth=req.breadth,
                depth=req.depth,
                quantity=req.quantity,
            )
        )

    if not entries:
        raise HTTPException(status_code=400, detail="No measurement entries provided")

    # ---------- Compute total quantity ----------
    total_qty = 0.0
    normalized_entries = []
    for e in entries:
        no_val = e.no_of_items or 0
        b_val = e.breadth or 0
        d_val = e.depth or 0
        l_val = e.length or 0

        q_val = e.quantity
        if q_val is None:
            q_val = (no_val or 0) * (l_val or 0) * (b_val or 0) * (d_val or 0)

        total_qty += q_val

        normalized_entries.append(
            {
                "pile_description": e.pile_description or "",
                "no": no_val,
                "b": b_val,
                "d": d_val,
                "l": l_val,
                "q": q_val,
            }
        )

    # ---------- SSR & BOQ lookups ----------
    ssr_item_no = ""
    unit_str = ""
    non_ssr = False

    if total_qty > 0:
        ssr_info = fetch_ssr_rate(req.description, total_qty)
        if ssr_info is None:
            non_ssr = True
        else:
            ssr_item_no = ssr_info.get("ssr_item_no", "") or ""
            unit_str = ssr_info.get("unit", "") or ""
    else:
        non_ssr = True

    boq_item_no = fetch_boq_item_no(req.description) or ""

    # ---------- PDF build ----------
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 40
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, y, "Item Measurement Sheet")
    y -= 18

    # Item description
    p.setFont("Helvetica", 9)
    header_text = f"Item Description (from SSR / BOQ): {req.description}"
    header_lines = wrap(header_text, 110)
    for line in header_lines:
        p.drawString(40, y, line)
        y -= 12

    y -= 4
    p.setFont("Helvetica-Bold", 9)
    if non_ssr:
        p.drawString(40, y, "SSR Item No: NON SSR ITEM")
    else:
        p.drawString(40, y, f"SSR Item No: {ssr_item_no or '-'}")
    y -= 12
    p.drawString(40, y, f"BOQ Item No: {boq_item_no or '-'}")
    y -= 16

    # Table header
    p.setFont("Helvetica-Bold", 9)
    p.drawString(40,  y, "Sr.")
    p.drawString(60,  y, "Pile Description")
    p.drawString(240, y, "No")
    p.drawString(285, y, "B")
    p.drawString(325, y, "D")
    p.drawString(365, y, "L")
    p.drawString(405, y, "Unit")
    p.drawString(460, y, "Quantity")
    y -= 12
    p.line(40, y, 560, y)
    y -= 8

    p.setFont("Helvetica", 8)
    line_height = 12

    for idx, row in enumerate(normalized_entries, start=1):
        if y - line_height < 60:
            p.showPage()
            y = height - 80
            p.setFont("Helvetica-Bold", 9)
            p.drawString(40,  y, "Sr.")
            p.drawString(60,  y, "Pile Description")
            p.drawString(240, y, "No")
            p.drawString(285, y, "B")
            p.drawString(325, y, "D")
            p.drawString(365, y, "L")
            p.drawString(405, y, "Unit")
            p.drawString(460, y, "Quantity")
            y -= 12
            p.line(40, y, 560, y)
            y -= 8
            p.setFont("Helvetica", 8)

        pile_desc = row["pile_description"].replace("\n", " ").strip()
        if len(pile_desc) > 35:
            pile_desc = pile_desc[:35] + "..."

        p.drawString(40, y, str(idx))
        p.drawString(60, y, pile_desc)
        p.drawRightString(270, y, f"{row['no']:.3f}")
        p.drawRightString(310, y, f"{row['b']:.3f}")
        p.drawRightString(350, y, f"{row['d']:.3f}")
        p.drawRightString(390, y, f"{row['l']:.3f}")

        p.drawString(405, y, unit_str if not non_ssr else "")

        p.drawRightString(540, y, f"{row['q']:.3f}")

        y -= line_height

    # Total
    if y < 60:
        p.showPage()
        y = height - 80

    p.setFont("Helvetica-Bold", 9)
    p.line(380, y, 560, y)
    y -= 12
    p.drawRightString(500, y, "Total Quantity:")
    p.drawRightString(560, y, f"{total_qty:.3f}")

    p.showPage()
    p.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=material_measurement_sheet.pdf"
        },
    )

# ============================================================
#  SINGLE MATERIAL BILL EXCEL SHEET 
# ============================================================

@app.post("/materials/single-bill/excel")
def download_single_material_bill_excel(
    req: schemas.SingleMaterialBillRequest,
):
    """
    Generate Excel for a single material's measurement sheet.
    Uses the same request as the PDF:
      - description: SSR/BOQ item description
      - entries: list of pile/measurement rows
    """

    if not req.entries:
        raise HTTPException(status_code=400, detail="No measurement entries provided")

    # Recompute quantity row-wise as No × L × B × D (same as frontend)
    rows = []
    for idx, e in enumerate(req.entries, start=1):
        no_ = float(e.no_of_items or 0)
        L = float(e.length or 0)
        B = float(e.breadth or 0)
        D = float(e.depth or 0)
        qty = no_ * L * B * D
        rows.append(
            {
                "sr": idx,
                "pile_description": e.pile_description or "",
                "no_of_items": no_,
                "length": L,
                "breadth": B,
                "depth": D,
                "quantity": qty,
            }
        )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Measurement"

    # Header: item description on top
    ws["A1"] = "Item Description:"
    ws["B1"] = req.description

    # Table header (start at row 3)
    header_row = 3
    headers = [
        "Sr. No.",
        "Pile Description",
        "No",
        "Length",
        "Breadth",
        "Depth",
        "Quantity",
    ]
    for col_idx, title in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=title)

    # Data rows (from row 4 onwards)
    excel_row = header_row + 1
    for r in rows:
        ws.cell(row=excel_row, column=1, value=r["sr"])
        ws.cell(row=excel_row, column=2, value=r["pile_description"])
        ws.cell(row=excel_row, column=3, value=r["no_of_items"])
        ws.cell(row=excel_row, column=4, value=r["length"])
        ws.cell(row=excel_row, column=5, value=r["breadth"])
        ws.cell(row=excel_row, column=6, value=r["depth"])
        ws.cell(row=excel_row, column=7, value=r["quantity"])
        excel_row += 1

    # Autosize a bit (simple version)
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="single_material_bill.xlsx"'
        },
    )